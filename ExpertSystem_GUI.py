#!/usr/bin/python3
# Coding: utf-8
# Author: Rogen
# Description: 專家系統主介面



from tkinter import *
from tkinter import ttk, messagebox, font, filedialog
from PIL import ImageTk, Image
from functools import partial
from openpyxl import load_workbook
from ExpertSystem_Functions import ExpertSystemFunctions
# from ExpertSystem_Functions import *
import os, subprocess, re
import pandas as pd
import shutil


class LanguageImport(object):
	''' For language selection '''
	def __init__(self, master, ver, sheet, code):
		super(LanguageImport, self).__init__()
		self.master = master
		self.ver = ver
		self.sheet = sheet
		self.code = code

	def Language(self):
		msgans = messagebox.askyesno('Question','TempFile will delete if you change language!!\nDo you want to change?', icon='warning')
		print(msgans)
		if msgans:
			from Language_Session import Language_Session
			self.master.destroy()

			ExpertSystemFunctions(self.ver,self.sheet,self.code).GuiClose()		# Remove photo of TempFile and update SitEva.csv
			LS = Language_Session()
			LS.main()
		else:
			pass


class ExpertSystemInterface(ExpertSystemFunctions):
	''' Expert system GUI interface '''
	def __init__(self, ver, sheet, code):
		super(ExpertSystemInterface, self).__init__(ver, sheet, code)
		self.ver = ver
		self.sheet = sheet
		self.code = code
		self.node_list = list(set(self.df.Node.tolist() + [int(re.search(r'#(\d+):',i).group(1)) 
			for i in self.df.Yes.dropna().tolist()+self.df.No.dropna().tolist() if re.search(r'#(\d+):',i)]))
		self.node_list.sort()


	# 診斷結果介面
	def done_window(self):
		def tree_build(event,frame):
			# 超參數設定
			# w, h = event.width, event.height
			# wid_ratio = self.screen_width/1920
			# hei_ratio = self.screen_height/1080
			wid_ratio = 1
			hei_ratio = 1
			bpadx = int(10*wid_ratio)
			
			# 表格框架設定
			tree_frame = Frame(frame)
			tree_frame.place(x=5, y=5, width=int(890*wid_ratio), height=int(510*hei_ratio))
			tree = ttk.Treeview(tree_frame, columns=('c1','c2','c3','c4','c5','c6','c7'), show="headings")
			tree.column('c1', width=int(wid_ratio*50), anchor='center')
			tree.column('c2', width=int(wid_ratio*300), anchor='w')
			tree.column('c3', width=int(wid_ratio*70), anchor='center')
			tree.column('c4', width=int(wid_ratio*70), anchor='center')
			tree.column('c5', width=int(wid_ratio*70), anchor='center')
			tree.column('c6', width=int(wid_ratio*500), anchor='center')
			tree.column('c7', width=int(wid_ratio*500), anchor='center')
			tree.tag_configure("main_branch",background='light sky blue')
			tree.tag_configure("secondary_branch",background='yellow')

			# 表格標題設定
			d = {'c1':self.interface['done_title']['node'],'c2':self.interface['done_title']['question'],
			'c3':self.interface['done_title']['answer'],'c4':self.interface['done_title']['yescore'],
			'c5':self.interface['done_title']['noscore'],'c6':self.interface['done_title']['diagnosis'],
			'c7':self.interface['done_title']['solution']}

			for header,name in d.items():
				tree.heading(header, text=name)

			# 表格x, y軸設定
			fr_y = Frame(tree_frame)
			fr_y.pack(side='right', fill='y')
			Bricks = Label(fr_y, borderwidth=1, relief='raised').pack(side='bottom', fill='x')
			sb_y = Scrollbar(fr_y, orient="vertical", command=tree.yview)
			sb_y.pack(side='right',fill='y')
			fr_x = Frame(tree_frame)
			fr_x.pack(side='bottom', fill='x')
			sb_x = Scrollbar(fr_x, orient="horizontal", command=tree.xview)
			sb_x.pack(side='bottom',fill='x')
			tree.config(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
			tree.pack(side='left',fill='both')


			# 按鈕功能設定
			try:
				sheet_index = self.rulebase_list[self.rulebase_list.index(self.sheet_name)+1]
			except IndexError as e:
				sheet_index = ''

			self.rulebase = StringVar()
			button_frame = Frame(frame)
			button_frame.place(x=int(200*wid_ratio), y=int(530*hei_ratio), width=int(880*wid_ratio), height=int(40*hei_ratio))

			self.save = Button(button_frame, text='Save', width=int(10*wid_ratio), height=1, command=self.save_diagnosis)
			self.next_rulebase = ttk.OptionMenu(button_frame, self.rulebase, sheet_index, *self.rulebase_list)
			self.submit = Button(button_frame, text='Next', width=int(10*wid_ratio), command=self.next_rulesbase_diag)
			# self.show = Button(button_frame, text='Show', width=int(10*wid_ratio), command=self.table4result)
			self.show = Button(button_frame, text='Show', width=int(10*wid_ratio), command=self.open_temp_images)
			self.showinexcel = Button(button_frame, text='Show in Excel', width=int(12*wid_ratio), state=ACTIVE, command=self.open_csv_excel)
			style = ttk.Style()
			style.configure('TMenubutton', foreground="black", background="white", width=int(15*wid_ratio), height=7, relief="raised")

			self.showinexcel.grid(row=0,column=0,padx=bpadx)
			self.show.grid(row=0,column=1,padx=bpadx)
			self.save.grid(row=0,column=2,padx=bpadx)
			self.next_rulebase.grid(row=0,column=3,padx=bpadx)
			self.submit.grid(row=0,column=4,padx=bpadx)
			self.diagnosis_done(tree)

		try:
			if self.done_win.state() == 'normal':
				self.tree_iterater = 0
				self.done_win.destroy()
		except:
			self.done_win = Toplevel(self.windows)
			# self.done_win.columnconfigure(0, weight=1)	 #control the column of windows Zoom 
			# self.done_win.rowconfigure(0, weight=1)		#control the row of windows Zoom 
			self.done_win.resizable(width=False, height=False)
			self.done_win.title(self.interface['windows']['done'])
			self.done_win.geometry('900x570')
			main_frame = Frame(self.done_win)
			main_frame.pack(fill=BOTH, expand=1)
			main_frame.bind('<Configure>',lambda event,x=main_frame: tree_build(event,x))


	# 增加規則 (不開放使用)
	def add_rule(self):
		def reset():
			newEQuery.delete(1.0, END)
			newCQuery.delete(1.0, END)
			newNode.set('')
			oldNode.set('')
			yesELink.set('')
			yesCLink.set('')
			noELink.set('')
			noCLink.set('')
			yesConf.set('')
			noConf.set('')
			imgNames.set('')
			noteLink.set(None)

		def save(new_rule, old_rule):
			count = 1
			rule_list = []
			book = load_workbook(self.rules_database_path, keep_vba=True)
			wb = book.get_sheet_by_name(self.sheet_name)

			for i in range(len(new_rule)):
				if i == 1 or i == 6:
					rule_list.append(new_rule[i].get(1.0,END).replace('\n',''))
				elif i == 0 or i == 4 or i == 5:
					if new_rule[i].get():
						try:
							node = int(new_rule[i].get())
						except ValueError:
							node = float(new_rule[i].get())
					else:
						node = new_rule[i].get()
					rule_list.append(node)
				else:
					rule_list.append(new_rule[i].get())
			
			rule_list.insert(6,'N'+new_rule[0].get())   #add photo folder
			rule_list.insert(4, '')					 #add goto node
			rules_num = len(self.df.axes[0])
			wb.append([rules_num] + rule_list)
			warning = check(rule_list, old_rule)

			if len(warning) > 0:
				messagebox.showwarning('WARNING', '\n'.join(warning))
			else:
				nodes = [i.value for i in wb['B']]
				ori_node = int(old_rule[0].get())

				if ori_node in nodes:
					if old_rule[1].get() == 'yes':
						wb.cell(row=nodes.index(ori_node)+1, column=4).value = '#' + str(rule_list[0])
					else:
						wb.cell(row=nodes.index(ori_node)+1, column=5).value = '#' + str(rule_list[0])

					if not os.path.exists(self.photo_path+'/'+rule_list[7]):
						os.makedirs(self.photo_path + '/' + rule_list[7], exist_ok=True)

					for i in self.img_list:
						img = Image.open(i)
						file_attention = i.split('.')[-1]
						img.save(self.photo_path + '/N' + str(rule_list[0]) + '/N' + str(rule_list[0]) + 'P' + str(count) + '.' + file_attention)
						count += 1

					book.save(self.rules_database_path)
					self.program_restart()
				else:
					messagebox.showerror('ERROR', 'The number of original node is not exist!')

		def check(new_data, old_data):
			warning_info = []
			if new_data[0] == '':
				warning_info.append('The "Node" number of new rule is empty!')
			if new_data[1] == '':
				warning_info.append('The "Question" in English is empty!')
			if new_data[8] == '':
				warning_info.append('The "Question" in Chinese is empty!')
			if new_data[2] == '':
				warning_info.append('The "Yes" statemant in English is empty!')
			if new_data[3] == '':
				warning_info.append('The "No" statemant in English is empty!')
			if (not bool(re.match(r'#\d+: Unknown', new_data[2]))) and bool(re.match(r'#\d+:.+', new_data[2])) and new_data[5] == '':
				warning_info.append('The "Yes confidence" is empty!')
			if (not bool(re.match(r'#\d+: Unknown', new_data[3]))) and bool(re.match(r'#\d+:.+', new_data[3])) and new_data[6] == '':
				warning_info.append('The "No confidence" is empty!')
			if bool(re.match(r'#\d+:.+', new_data[2])) and new_data[9] == '':
				warning_info.append('The "Yes" statement in Chinese is empty!')
			if bool(re.match(r'#\d+:.+', new_data[3])) and new_data[10] == '':
				warning_info.append('The "No" statement in Chinese is empty!')
			if old_data[0].get() == '':
				warning_info.append('The "Node" number will be linked to new rule is empty!')
			if old_data[1].get() == 'None':
				warning_info.append('The "Yes/NO" statement in the node which will linked is empty!')
			return warning_info

		def choose_file():
			self.img_list = []
			root = Tk()
			root.withdraw()
			SelectFilename = filedialog.askopenfilenames(initialdir = "/", title='Select Input image', filetype=[("PNG",".png"),("GIF",".gif"),("JPG",".jpg"),("all file","*.*")])
			self.img_list =  self.img_list + list(SelectFilename)
			imgNames.set([i.split('/')[-1] for i in self.img_list])

		self.add_win = Toplevel(self.windows)
		self.add_win.grab_set() #lock on the windows in the progress, but it's not always on the top.
		# self.add_win.attributes('-topmost',True) #lock on the windows in the progress, and it's topmost.
		self.add_win.title('Add Rule')
		self.add_win.resizable(width=False, height=False)
		self.add_win.geometry('600x450')
		
		self.img_list = []
		newNode = StringVar(value='')
		oldNode = StringVar(value='')
		yesELink = StringVar(value='')
		yesCLink = StringVar(value='')
		noELink = StringVar(value='')
		noCLink = StringVar(value='')
		yesConf = StringVar(value='')
		noConf = StringVar(value='')
		imgNames = StringVar(value='')
		noteLink = StringVar()
		noteLink.set(None)

		Label(self.add_win, text='(Please Answer the Query as Follows for Adding New Rule)').place(x=30,y=20)
		Label(self.add_win, text='Node: ').place(x=100,y=50)
		new_node = Entry(self.add_win, textvariable=newNode, width=7).place(x=150,y=50)
		Label(self.add_win, text='Question in English: ').place(x=30,y=80)
		newEQuery = Text(self.add_win, height=2, width=50)
		newEQuery.place(x=150,y=80)
		Label(self.add_win, text='Question in Chinese: ').place(x=30,y=130)
		newCQuery = Text(self.add_win, height=2, width=50)
		newCQuery.place(x=150,y=130)

		Label(self.add_win, text='(Please Input the Node Connected / Diagnosed Result in Below)').place(x=30,y=170)
		Label(self.add_win, text='The "YES" Pathway (Fill in English): ').place(x=30,y=200)
		yes_elink = Entry(self.add_win, textvariable=yesELink).place(x=250,y=200)
		Label(self.add_win, text='The "YES" Pathway (Fill in Chinese): ').place(x=30,y=230)
		yes_clink = Entry(self.add_win, textvariable=yesCLink).place(x=250,y=230)

		Label(self.add_win, text='The "NO" Pathway (Fill in English): ').place(x=30,y=260)
		no_elink = Entry(self.add_win, textvariable=noELink).place(x=250,y=260)
		Label(self.add_win, text='The "NO" Pathway (Fill in Chinese): ').place(x=30,y=290)
		no_clink = Entry(self.add_win, textvariable=noCLink).place(x=250,y=290)

		Label(self.add_win, text='Yes Conf.(0~1):').place(x=410,y=215)
		yes_conf = Entry(self.add_win, textvariable=yesConf, width=7).place(x=510,y=215)
		Label(self.add_win, text='No Conf.(0~1):').place(x=410,y=275)
		no_conf = Entry(self.add_win, textvariable=noConf, width=7).place(x=510,y=275)
		Label(self.add_win, textvariable=imgNames, width=60).place(x=80, y=315)

		Canvas(self.add_win, width=600, height=2, highlightthickness=0, bg="black").place(x=0,y=350)

		Label(self.add_win, text='Which node will link to new rule: ').place(x=30,y=360)
		old_node = Entry(self.add_win, textvariable=oldNode, width=7).place(x=230,y=360)
		Label(self.add_win, text='Which statemant will link: ').place(x=30,y=390)
		rYES = Radiobutton(self.add_win, text='YES', var=noteLink, value='yes').place(x=200,y=390)
		rNO = Radiobutton(self.add_win, text='NO', var=noteLink, value='no').place(x=250,y=390)

		new_rule = [newNode, newEQuery, yesELink, noELink, yesConf, noConf, newCQuery, yesCLink, noCLink]
		old_rule = [oldNode, noteLink]

		Button(self.add_win, text='View Rules Structure', height=1, width=20, command=partial(super().__graph__, self.add_win)).place(x=430, y=15)
		Button(self.add_win, text='Reset', height=1, width=10, command=reset).place(x=380, y=400)
		Button(self.add_win, text='Save', height=1, width=10, command=partial(save, new_rule, old_rule)).place(x=470, y=400)
		Button(self.add_win, text='Upload Images to Annotated Question', height=1, width=32, command=choose_file).place(x=10, y=315)


	# 刪除規則 (不開放使用)
	def delete_rule(self):
		def search_link():
			separate = 0
			influ_nodes_by_yes = [self.df.Node[i] for i in range(len(self.df.index)) if delNum.get() == self.df.Yes[i].split(':')[0].replace('#','')]
			influ_nodes_by_no = [self.df.Node[i] for i in range(len(self.df.index)) if delNum.get() == self.df.No[i].split(':')[0].replace('#','')]

			if len(influ_nodes_by_yes) == 0 and len(influ_nodes_by_no) == 0: #not find affect node
				Label(self.del_win, text='').place(x=30,y=100)

			for y in influ_nodes_by_yes: #find affect node is yes
				Label(self.del_win, text='Node '+str(y)+' Connected with Deleting Node by "Yes", Please retrieve Connected Site: ').place(x=30,y=70+separate)
				Entry(self.del_win, text='', width=12).place(x=30,y=100+separate)
				separate += 50

			for n in influ_nodes_by_no: #find affect node is no
				Label(self.del_win, text='Node '+str(n)+' Connected with Deleting Node by "No", Please retrieve Connected Site: ').place(x=30,y=70+separate)
				Entry(self.del_win, text='', width=12).place(x=30,y=100+separate)
				separate += 50

			# l.update_idletasks(); e.update_idletasks()

		def del_note():
			book = load_workbook(self.rules_database_path, keep_vba=True)
			wb = book.get_sheet_by_name(self.sheet_name)
			yesDiag, noDiag = self.diagnosis_node_index(delNum)

			try:
				row = [col.value for col in wb['B']].index(delNum.get())+1
				tag = 1
			except:
				if len(yesDiag):
					row = [col.value for col in wb['D']].index(yesDiag[0])+1
					tag = 2
				else:
					row = [col.value for col in wb['E']].index(noDiag[0])+1
					tag = 3

			if tag == 1:
				for r in range(row, wb.max_row):
					for c in range(1,wb.max_column+1):
						wb.cell(r,c).value = wb.cell(r+1,c).value
				for cell in list(wb.rows)[wb.max_row-1]:
					cell.value = None
			elif tag == 2:
				wb.cell(row,4).value = None
			elif tag == 3:
				wb.cell(row,5).value = None

			book.save(self.rules_database_path)
			shutil.rmtree(self.photo_path+'/'+str(delNum.get()))
			self.program_restart()

		delNum = IntVar()
		self.del_win = Toplevel(self.windows)
		self.del_win.grab_set()
		self.del_win.title('Delete Rule')
		self.del_win.resizable(width=False, height=False)
		self.del_win.geometry('500x150')
		Label(self.del_win, text='Which note will be delete?').place(x=30,y=30)
		# entry = Entry(self.del_win, textvariable=delNum, width=12)
		# entry.config(fg='gray')
		# entry.bind('<FocusIn>', lambda event, objects=entry, show='Example: #7': self.on_entry_click(objects, show))
		# entry.bind('<FocusOut>', lambda event, objects=entry, show='Example: #7': self.on_focusout(objects, show))
		# entry.place(x=170,y=20)
		delNode = ttk.Combobox(self.del_win, width=7, textvariable=delNum, values=self.node_list)
		delNode.current(0)
		delNode.place(x=200,y=30)
		# Button(self.del_win, text='Search Influenced Node',command=search_link).place(x=300,y=30)
		Button(self.del_win, text='Submit',width=7, command=del_note).place(x=280,y=30)
		Button(self.del_win, text='View Rules Structure',command=partial(super().__graph__, self.del_win)).place(x=350,y=30)


	# 修改規則 (不開放使用)
	def edit_rule(self):
		def menu():
			count = 0
			edit_submit.config(state='active')

			if not editNode.get():
				messagebox.showerror('ERROR', 'Please input which node want to retrieve.')
			elif editNode.get() in self.df.Node.tolist():		 #node rules
				clear_menu(self.leaf)
				editList = list(self.df.Node == editNode.get())
				yesConf.set(self.df.Yes_score[editList].tolist()[0])
				noConf.set(self.df.No_score[editList].tolist()[0])

				l_l1 = Label(self.edit_win, text='Question in English:')
				l_t1 = Text(self.edit_win, height=2, width=20)
				l_l2 = Label(self.edit_win, text='Question in Chinese:')
				l_t2 = Text(self.edit_win, height=2, width=20)
				l_t1.insert(1.0,self.df.Question[editList].tolist()[0])
				l_t2.insert(1.0,self.df.中文問題[editList].tolist()[0])

				englishList = list(set(['#'+str(i) for i in self.df.Node.dropna().tolist()]+self.df.Yes.dropna().tolist()+self.df.No.tolist()))
				englishList.remove('Max Possible')
				englishList.sort()

				l_l3 = Label(self.edit_win, text='Connecting "Yes" Path to:')
				l_c3 = ttk.Combobox(self.edit_win,state='readonly',width=15,values=englishList)
				l_c3.current(englishList.index(self.df.Yes[self.df.Node == editNode.get()].tolist()[0]))
				l_l4 = Label(self.edit_win, text='Connecting "No" Path to:')
				l_c4 = ttk.Combobox(self.edit_win,state='readonly',width=15,values=englishList)
				l_c4.current(englishList.index(self.df.No[self.df.Node == editNode.get()].tolist()[0]))

				chineseList = list(set(self.df.Yes診斷.dropna().tolist() + self.df.No診斷.dropna().tolist() + ['']))
				chineseList.sort()

				l_l5 = Label(self.edit_win, text='"Yes" Path in Chinese:')
				l_c5 = ttk.Combobox(self.edit_win,state='readonly',width=15,values=chineseList)
				if str(self.df.Yes診斷[self.df.Node == editNode.get()].tolist()[0]) == 'nan':
					l_c6.current(0)
				else:
					l_c5.current(chineseList.index(self.df.Yes診斷[self.df.Node == editNode.get()].tolist()[0]))
				l_l6 = Label(self.edit_win, text='"No" Path in Chinese:')
				l_c6 = ttk.Combobox(self.edit_win,state='readonly',width=15,values=chineseList)
				if str(self.df.No診斷[self.df.Node == editNode.get()].tolist()[0]) == 'nan':
					l_c6.current(0)
				else:
					l_c6.current(chineseList.index(self.df.No診斷[self.df.Node == editNode.get()].tolist()[0]))

				l_l7 = Label(self.edit_win, text='Confidence of Yes:')
				l_e7 = Entry(self.edit_win, width=7, textvariable=yesConf)
				l_l8 = Label(self.edit_win, text='Confidence of No:')
				l_e8 = Entry(self.edit_win, width=7, textvariable=noConf)

				self.node = [l_l1,l_t1,l_l2,l_t2,l_l3,l_c3,l_l4,l_c4,l_l5,l_c5,l_l6,l_c6,l_l7,l_e7,l_l8,l_e8]
				coordinate_x = [30,200,30,200,30,200,30,200,30,200,30,200,350,470,350,470]
				coordinate_y = [100,100,140,140,180,180,210,210,240,240,270,270,200,200,260,260]
				for j in self.node:
					j.place(x=coordinate_x[count],y=coordinate_y[count])
					count += 1
			elif editNode.get() <= self.node_list[-1]:	#leaf rules
				clear_menu(self.node)
				n_l1 = Label(self.edit_win, text='Diagnosis in English:')
				n_t1 = Text(self.edit_win, height=2, width=20)
				n_l2 = Label(self.edit_win, text='Diagnosis in Chinese:')
				n_t2 = Text(self.edit_win, height=2, width=20)
				yesDiag, noDiag = self.diagnosis_node_index(editNode)

				if len(yesDiag):
					text1 = yesDiag[0]
					text2 = self.df.Yes診斷[self.df.Yes.tolist().index(yesDiag[0])]
				else:
					text1 = noDiag[0]
					text2 = self.df.No診斷[self.df.No.tolist().index(noDiag[0])]
				n_t1.insert(1.0,text1)
				n_t2.insert(1.0,text2)

				self.leaf = [n_l1,n_t1,n_l2,n_t2]
				coordinate_x = [30,250,30,250]
				coordinate_y = [100,100,140,140]
				for i in self.leaf:
					i.place(x=coordinate_x[count],y=coordinate_y[count])
					count += 1
			else:
				messagebox.showwarning('WARNING','Not find the node.')

		def clear_menu(objects):
			for o in objects:	
				try:
					o.destroy()
				except:
					print(o,'is not exist!')

		def submit():
			book = load_workbook(self.rules_database_path, keep_vba=True)
			wb = book.get_sheet_by_name(self.sheet_name)
			tag = 0
			try:
				row = [col.value for col in wb['B']].index(editNode.get())+1
				tag = 1
			except:
				yesDiag, noDiag = self.diagnosis_node_index(editNode)
				if len(yesDiag):
					row = [col.value for col in wb['D']].index(yesDiag[0])+1
					tag = 2
				else:
					row = [col.value for col in wb['E']].index(noDiag[0])+1
					tag = 3
			if tag == 1:
				wb.cell(row,3).value = self.node[1].get(1.0,END).replace('\n','')
				wb.cell(row,4).value = self.node[5].get()
				wb.cell(row,5).value = self.node[7].get()
				wb.cell(row,7).value = yesConf.get()
				wb.cell(row,8).value = noConf.get()
				wb.cell(row,10).value = self.node[3].get(1.0,END).replace('\n','')
				wb.cell(row,11).value = self.node[9].get()
				wb.cell(row,12).value = self.node[11].get()
			elif tag == 2:
				wb.cell(row,4).value = self.leaf[1].get(1.0,END).replace('\n','')
				wb.cell(row,11).value = self.leaf[3].get(1.0,END).replace('\n','')
			elif tag == 3:
				wb.cell(row,5).value = self.leaf[1].get(1.0,END).replace('\n','')
				wb.cell(row,12).value = self.leaf[3].get(1.0,END).replace('\n','')

			book.save(self.rules_database_path)
			self.program_restart()

		self.node = []; self.leaf = []
		editNode = IntVar()
		yesConf = StringVar()
		noConf = StringVar()
		self.edit_win = Toplevel(self.windows)
		self.edit_win.grab_set()
		self.edit_win.title('Edit Rule')
		self.edit_win.geometry('550x400')
		self.edit_win.resizable(False,False)
		Label(self.edit_win, text='Which note will be edit?').place(x=30,y=40)
		nodeBox = ttk.Combobox(self.edit_win, width=7, textvariable=editNode, state='readonly', values=self.node_list)
		nodeBox.current(0)
		nodeBox.place(x=180,y=40)
		Button(self.edit_win, text='View Rules Structure', width=20, command=partial(super().__graph__, self.edit_win)).place(x=280,y=40)
		Button(self.edit_win, text='Submit', width=10, command=menu).place(x=450,y=40)
		Canvas(self.edit_win, width=600, height=2, bg="black").place(x=0,y=80)
		Canvas(self.edit_win, width=600, height=2, bg="black").place(x=0,y=320)
		edit_submit = Button(self.edit_win, text='Submit', width=10, command=submit, state='disabled')
		edit_submit.place(x=400,y=350)


	# 照片呈現介面
	def photo_window(self,master=None,image=None):
		# print(self.photo_images[self.photo_image_counter])
		# print(image)
		if image == None:
			image = self.im
		if master == None:
			master = self.windows

		if not re.search('img_not_available.png',self.photo_images[0]): #if no image, don't open new window
			try:
				if self.photo_win.state() == 'normal':
					self.photo_win.destroy()
			except:
				self.photo_win = Toplevel(master)
				self.photo_win.title('Figure magnifier')
				self.photo_win.geometry('%dx%d' % (round(1200*self.screen_width/1920), round(900*self.screen_height/1080)))
				self.figure_magnification(image)


	# 詢問規則路徑介面
	def rulepath_window(self):
		try:
			if self.rulepath_win.state() == 'normal':
				pass
		except:
			# ---------- option list ----------
			self.rulepath_win = Toplevel(self.windows)
			self.rulepath_win.title('Rule path Rocode')
			self.rulepath_win.geometry('500x500')
			self.rulepath_win.resizable(width=False, height=False)

			self.optiontree = ttk.Treeview(self.rulepath_win, height=9, columns=('c1','c2','c3'), show="headings")
			self.optiontree.column('c1', width=70, anchor=CENTER)
			self.optiontree.column('c2', width=300, anchor=W)
			self.optiontree.column('c3', width=100, anchor=CENTER)

			tableheads = {'c1':self.interface['done_title']['node'],
						 'c2':self.interface['done_title']['question'],
						 'c3':self.interface['done_title']['answer']}

			for header,name in tableheads.items():
				 self.optiontree.heading(header, text=name)
				
			op_y_axial = Frame(self.optiontree)
			op_y_axial.pack(side='right', fill='y')
			Label(op_y_axial, borderwidth=1, relief='raised', font="Arial 10").pack(side='bottom', fill='x')
			op_y_sb = Scrollbar(op_y_axial, orient="vertical", command=self.optiontree.yview)
			op_y_sb.pack(side='right',fill='y')
			
			op_x_axial = Frame(self.optiontree)
			op_x_axial.pack(side='bottom', fill='x')
			op_x_sb = Scrollbar(op_x_axial, orient="horizontal", command=self.optiontree.xview)
			op_x_sb.pack(side='bottom',fill='x')
			
			self.optiontree.configure(yscrollcommand=op_y_sb.set, xscrollcommand=op_x_sb.set)
			self.optiontree.pack(side=LEFT, fill=BOTH, expand=True)

			self.option_record('showhand')


	# 視窗關閉返回主選單
	def wm_delete_window(self):
		from Option_List import optionlist
		# self.GuiClose()
		self.windows.destroy()
		manu = optionlist(self.language)
		manu.main()


	# 專家系統主介面
	def gui(self):
		self.windows = Tk()
		self.windows.resizable(width=False, height=False)
		self.windows.title(self.interface['windows']['main'])
		self.windows.protocol("WM_DELETE_WINDOW",self.wm_delete_window)

		self.screen_width = self.windows.winfo_screenwidth()
		self.screen_height = self.windows.winfo_screenheight()

		w_ratio = self.screen_width/1920
		h_ratio = self.screen_height/1080
		# w_ratio = 1
		# h_ratio = 1
		win_width = round(1140*w_ratio)
		win_height = round(755*h_ratio)
		win_xanchor = 200
		win_yanchor = 100
		self.windows.geometry('{}x{}+{}+{}'.format(win_width,win_height,win_xanchor,win_yanchor))

		#------- create menubar ----------
		menubar = Menu(self.windows)

		filemenu = Menu(menubar, tearoff=0)
		menubar.add_cascade(label=self.interface['menubar']['function'],menu=filemenu)
		filemenu.add_command(label=self.interface['menubar']['restart'],command=self.program_restart)
		filemenu.add_command(label=self.interface['menubar']['language'],command=LanguageImport(self.windows,self.ver,self.sheet,self.code).Language)
		filemenu.add_separator()
		filemenu.add_command(label=self.interface['menubar']['exit'],command=self.wm_delete_window)

		editmenu = Menu(menubar, tearoff=0)
		menubar.add_cascade(label=self.interface['menubar']['edit'],menu=editmenu)
		editmenu.add_command(label=self.interface['menubar']['add_rule'],command=self.add_rule)
		editmenu.add_command(label=self.interface['menubar']['delete_rule'],command=self.delete_rule)
		editmenu.add_command(label=self.interface['menubar']['edit_rule'],command=self.edit_rule)

		# submenu = Menu(editmenu, tearoff=0)
		# editmenu.add_cascade(label=self.interface['menubar']['rulebase'], menu=submenu, underline=0)
		# submenu.add_cascade(label=self.interface['menubar']['import'], underline=0)

		exportmenu = Menu(menubar, tearoff=0)
		menubar.add_cascade(label=self.interface['menubar']['tool'],menu=exportmenu)
		exportmenu.add_command(label=self.interface['menubar']['importresult'],command=self.save_import)
		exportmenu.add_command(label=self.interface['menubar']['diagnosis'],command=self.diagnosis_export)
		exportmenu.add_command(label=self.interface['menubar']['rulepath'],command=self.rulepath_window)
		exportmenu.add_command(label=self.interface['menubar']['satellite'],command=self.open_temp_images)

		self.windows.config(menu=menubar)

		#------- query label ----------
		font_size = round(14*w_ratio)
		querytitle = self.query_title(self.cur_node)

		self.queryframe = LabelFrame(self.windows, fg='blue', font=('Arial', font_size), text=querytitle)
		self.queryframe.place(x=round(15*w_ratio),y=round(10*h_ratio))
		self.querylabel = Label(self.queryframe, text=self.query, font=('Arial', 14), wraplength=round(690*w_ratio), height=round(7*h_ratio), width=round(65*w_ratio), bg='lightgray')
		self.querylabel.pack(anchor=W)

		#---------- option ----------
		answerframe = LabelFrame(self.windows, fg='blue', font=('Arial', font_size), text=self.interface['main_title']['answer'])
		answerframe.place(x=round(750*w_ratio),y=round(10*h_ratio))

		answer_top = Frame(answerframe)
		answer_top.pack()
		rpadx = round(10*w_ratio)
		rpady = round(22*h_ratio)

		self.user_answer = StringVar()
		self.user_answer.set(None)
		self.r1 = Radiobutton(answer_top, text=self.interface['radiobutten']['yes'], font=('Arial', font_size), variable=self.user_answer, value='yes', command=lambda x=1: self.analyze_button_control(x))
		self.r1.grid(row=1, column=1, padx=rpadx, pady=rpady, ipadx=0, ipady=0)
		self.r2 = Radiobutton(answer_top, text=self.interface['radiobutten']['no'], font=('Arial', font_size), variable=self.user_answer, value='no', command=lambda x=2: self.analyze_button_control(x))
		self.r2.grid(row=1, column=2, padx=rpadx, pady=rpady, ipadx=0, ipady=0)
		self.r3 = Radiobutton(answer_top, text=self.interface['radiobutten']['unknown'], font=('Arial', font_size), variable=self.user_answer, value='unknown',state=DISABLED)
		self.r3.grid(row=1, column=3, padx=rpadx, pady=rpady, ipadx=0, ipady=0)

		#---------- submit ----------
		answer_bottom = Frame(answerframe)
		answer_bottom.pack()
		bpadx = round(10*w_ratio)
		bpady = round(22*h_ratio)
		bwidth = round(9*w_ratio)
		bheight = round(2*h_ratio)
		self.back_button = Button(answer_bottom, text=self.interface['butten']['back'], width=bwidth, height=bheight, state='disabled', command=self.back_step)
		self.back_button.grid(row=1, column=1, padx=bpadx, pady=bpady)
		self.next_button = Button(answer_bottom, text=self.interface['butten']['next'], width=bwidth, height=bheight, command=self.next_step)
		self.next_button.grid(row=1, column=2, padx=bpadx, pady=bpady)
		self.analyze_button = Button(answer_bottom, text=self.interface['butten']['analyze'], width=bwidth, height=bheight, state='disabled', command=self.external_link)
		self.analyze_button.grid(row=1, column=3, padx=bpadx, pady=bpady)
		self.submit_button = Button(answer_bottom, text=self.interface['butten']['done'], width=bwidth, height=bheight, state='disabled', command=self.done_window)
		# self.submit_button = Button(answer_bottom, text=self.interface['butten']['done'], width=bwidth, height=bheight, state='disabled', command=lambda x=self.windows: self.__table__(x))
		self.submit_button.grid(row=1, column=4, padx=bpadx, pady=bpady)

		#---------- figure ----------
		self.label_height = round(475*h_ratio)
		self.label_width = round(715*w_ratio)
		fpadx = round(10*w_ratio)
		fpady = round(5*h_ratio)
		mwidth = round(3*w_ratio)
		mheight = round(1*h_ratio)

		self.figframe = LabelFrame(self.windows, fg='blue', font=('Arial',font_size))
		self.figframe.place(x=round(15*w_ratio),y=round(200*h_ratio))

		fig_top = Frame(self.figframe)
		fig_top.pack()
		fig_bottom = Frame(self.figframe)
		fig_bottom.pack()

		im = self.pri_photo()
		self.fig_label = Label(fig_top, height=self.label_height, width=self.label_width, image=im)
		self.fig_label.pack(fill='both')
		
		photo_back_button = Button(fig_bottom, text=' ◀◀ ', width=mwidth, height=mheight, command=partial(self.figure_iterator,'back'))
		photo_back_button.grid(row=1,column=1,padx=fpadx,pady=fpady)
		magnifier = Button(fig_bottom, text='⚙️', width=mwidth, height=mheight, command=self.photo_window)
		magnifier.grid(row=1,column=2,padx=fpadx,pady=fpady)
		photo_forward_button = Button(fig_bottom, text=' ▶▶ ', width=mwidth, height=mheight, command=partial(self.figure_iterator,'forward'))
		photo_forward_button.grid(row=1,column=3,padx=fpadx,pady=fpady)

		#---------- Query description ----------
		query_descript_frame = LabelFrame(self.windows,text=self.interface['main_title']['query_descript'],fg='blue',font=('Arial',font_size))
		query_descript_frame.place(x=round(750*w_ratio),y=round(200*h_ratio))

		self.query_desc_text = Text(query_descript_frame, font=("Helvetica", 12), height=round(13*h_ratio), width=round(40*w_ratio))
		texture = self.df.問題說明[self.cur_node == self.df.Node].tolist()[0]
		if str(texture) == 'nan':
			self.query_desc_text.insert(1.0,'')
		else:
			self.query_desc_text.insert(1.0,texture)
		self.query_desc_text.pack(side='left')

		#---------- Figure description ----------
		fig_descript_frame = LabelFrame(self.windows,text=self.interface['main_title']['figure_descript'],fg='blue',font=('Arial',font_size))
		fig_descript_frame.place(x=round(750*w_ratio),y=round(475*h_ratio))

		self.fig_desc_text = Text(fig_descript_frame,font=("Helvetica", 12),height=round(13*h_ratio),width=round(40*w_ratio))
		texture = self.df.圖片說明[self.cur_node == self.df.Node].tolist()[0]
		if str(texture) == 'nan':
			self.fig_desc_text.insert(1.0,'')
		else:
			self.fig_desc_text.insert(1.0,texture)
		self.fig_desc_text.pack(side='left')

		self.pri_query()
		self.windows.mainloop()


# if __name__ == '__main__':
# 	a = ExpertSystemInterface(2,'SiteEvaluation','001-0001')
# 	a.gui()
